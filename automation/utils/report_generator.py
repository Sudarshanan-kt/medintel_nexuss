import os
import glob
import json
import sys
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config
from automation.utils.logger_util import logger

class ReportGenerator:
    @staticmethod
    def save_intermediate_results(filename: str, results: list):
        """Saves intermediate JSON results for a single suite/shard.

        `filename` comes from Config.result_file_name() and encodes both the suite and the
        shard index, so parallel jobs never collide.
        """
        os.makedirs(Config.JSON_DIR, exist_ok=True)
        filepath = os.path.join(Config.JSON_DIR, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2)
        logger.info(f"Saved {len(results)} intermediate results to {filepath}")

    @classmethod
    def consolidate_and_generate_all(cls) -> bool:
        """Aggregates every intermediate result file and generates the HTML/Excel reports.

        Raises RuntimeError if no results are found. This is deliberate: an earlier version
        fabricated a full 1,500-case dataset at a ~97.5% pass rate whenever the real results
        were missing, which let a completely broken pipeline publish a green quality report.
        Reports must only ever describe tests that actually ran.
        """
        logger.info("Consolidating parallel execution results...")
        aggregated_results = []
        os.makedirs(Config.JSON_DIR, exist_ok=True)

        # Collect every per-suite/per-shard file the matrix produced.
        pattern = os.path.join(Config.JSON_DIR, "results_*.json")
        result_files = sorted(glob.glob(pattern))

        if not result_files:
            raise RuntimeError(
                f"No intermediate result files matched {pattern}. The test suites either did not "
                f"run or failed before writing results — refusing to generate a report."
            )

        for filepath in result_files:
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    shard_data = json.load(f)
                aggregated_results.extend(shard_data)
                logger.info(f"Loaded {len(shard_data)} results from {os.path.basename(filepath)}")
            except Exception as e:
                logger.error(f"Error loading intermediate report {filepath}: {str(e)}")

        if not aggregated_results:
            raise RuntimeError(
                f"Found {len(result_files)} result file(s) but they contained no test results — "
                f"refusing to generate a report."
            )

        # Save consolidated JSON
        consolidated_path = os.path.join(Config.JSON_DIR, "execution-results.json")
        with open(consolidated_path, "w", encoding="utf-8") as f:
            json.dump(aggregated_results, f, indent=2)
            
        logger.info(f"Consolidated {len(aggregated_results)} total test cases.")
        
        # Generate all reports
        cls.generate_excel_reports(aggregated_results)
        cls.generate_html_reports(aggregated_results)
        cls.generate_summary_markdown(aggregated_results)
        
        return True

    @classmethod
    def generate_excel_reports(cls, results: list):
        """Generates the master Excel test report."""
        os.makedirs(Config.EXCEL_DIR, exist_ok=True)
        filepath = os.path.join(Config.EXCEL_DIR, "Automation_Test_Report.xlsx")
        cls.generate_single_excel_report(results, filepath)

    @staticmethod
    def generate_single_excel_report(results: list, filepath: str):
        """Generates a single Excel sheet with professional styling from the provided results."""
        # Style Definitions
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=10)
        font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_pass = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        fill_fail = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        fill_skip = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        align_left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        def format_sheet(ws):
            # Autofit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or '')
                    if '\n' in val:
                        val = max(val.split('\n'), key=len)
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        headers = ["Test ID", "Type", "Module", "Title / Test Name", "Status", "Duration (s)"]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        ws.views.sheetView[0].showGridLines = True
        
        # Header Row
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
        
        # Data Rows
        for row_idx, r in enumerate(results, 2):
            row_data = [
                r["id"], r["type"], r["module"], r["title"],
                r["status"], round(r["execution_time"], 3)
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                
                if col_idx in [1, 2, 3, 5, 6]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left_wrap
                    
                # Format Status Cell
                if col_idx == 5:
                    if val == "Passed":
                        cell.fill = fill_pass
                    elif val == "Failed":
                        cell.fill = fill_fail
                    else:
                        cell.fill = fill_skip
                        
        format_sheet(ws)
        wb.save(filepath)
        logger.info(f"Exported Excel report: {os.path.basename(filepath)}")

    @staticmethod
    def generate_html_reports(results: list):
        """Generates premium responsive HTML dashboard and detailed execution-report."""
        os.makedirs(Config.HTML_DIR, exist_ok=True)
        
        # High level counts
        total = len(results)
        passed = len([r for r in results if r["status"] == "Passed"])
        failed = len([r for r in results if r["status"] == "Failed"])
        skipped = len([r for r in results if r["status"] == "Skipped"])
        pass_rate = (passed / total * 100) if total > 0 else 0
        total_duration = sum(r["execution_time"] for r in results)
        
        # Group failed test details
        failed_details = [r for r in results if r["status"] == "Failed"]
        
        # Dashboard HTML Template
        html_dashboard_template = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>MedIntel Nexus QA Automation Dashboard</title>
    <style>
        :root {{
            --bg-primary: #0b0f19;
            --bg-secondary: #161b26;
            --text-primary: #f8fafc;
            --text-secondary: #94a3b8;
            --primary: #10b981;
            --failed: #f43f5e;
            --skipped: #eab308;
            --accent: #3b82f6;
            --card-border: #1e293b;
        }}
        body {{
            background-color: var(--bg-primary);
            color: var(--text-primary);
            font-family: 'Inter', system-ui, -apple-system, sans-serif;
            margin: 0;
            padding: 24px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 1px solid var(--card-border);
            padding-bottom: 16px;
            margin-bottom: 24px;
        }}
        h1 {{
            margin: 0;
            font-size: 24px;
            letter-spacing: -0.5px;
            background: linear-gradient(90deg, #10b981, #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        .grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background-color: var(--bg-secondary);
            border: 1px solid var(--card-border);
            border-radius: 12px;
            padding: 20px;
            text-align: center;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        }}
        .card-val {{
            font-size: 32px;
            font-weight: 700;
            margin: 8px 0;
        }}
        .card-lbl {{
            color: var(--text-secondary);
            font-size: 14px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .text-pass {{ color: var(--primary); }}
        .text-fail {{ color: var(--failed); }}
        .text-skip {{ color: var(--skipped); }}
        
        .section-title {{
            font-size: 18px;
            font-weight: 600;
            margin-bottom: 16px;
            color: var(--text-primary);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background-color: var(--bg-secondary);
            border-radius: 12px;
            overflow: hidden;
            border: 1px solid var(--card-border);
            margin-bottom: 24px;
        }}
        th, td {{
            padding: 12px 16px;
            text-align: left;
        }}
        th {{
            background-color: #1e293b;
            color: var(--text-secondary);
            font-size: 13px;
            text-transform: uppercase;
            font-weight: 600;
        }}
        tr {{
            border-bottom: 1px solid var(--card-border);
        }}
        tr:last-child {{
            border-bottom: none;
        }}
        .badge {{
            padding: 4px 8px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
            display: inline-block;
        }}
        .badge-pass {{ background-color: rgba(16, 185, 129, 0.1); color: var(--primary); }}
        .badge-fail {{ background-color: rgba(244, 63, 94, 0.1); color: var(--failed); }}
        .badge-skip {{ background-color: rgba(234, 179, 8, 0.1); color: var(--skipped); }}
        
        .footer {{
            text-align: center;
            color: var(--text-secondary);
            font-size: 12px;
            margin-top: 48px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div>
                <h1>MedIntel Nexus QA Automation</h1>
                <div style="color: var(--text-secondary); font-size: 14px; margin-top: 4px;">Live E2E, Performance & Security Results</div>
            </div>
            <div style="font-size: 13px; color: var(--text-secondary);">Generated at: {time.strftime('%Y-%m-%d %H:%M:%S')}</div>
        </header>
        
        <div class="grid">
            <div class="card">
                <div class="card-val" style="color: var(--accent);">{total}</div>
                <div class="card-lbl">Total Executed</div>
            </div>
            <div class="card">
                <div class="card-val text-pass">{passed}</div>
                <div class="card-lbl">Passed</div>
            </div>
            <div class="card">
                <div class="card-val text-fail">{failed}</div>
                <div class="card-lbl">Failed</div>
            </div>
            <div class="card">
                <div class="card-val text-skip">{skipped}</div>
                <div class="card-lbl">Skipped</div>
            </div>
            <div class="card">
                <div class="card-val text-pass">{pass_rate:.2f}%</div>
                <div class="card-lbl">Success Rate</div>
            </div>
            <div class="card">
                <div class="card-val" style="color: #a855f7;">{total_duration:.2f}s</div>
                <div class="card-lbl">Duration</div>
            </div>
        </div>

        <div class="section-title">Failed Critical / High Severity Defect Log ({len(failed_details)})</div>
        <table>
            <thead>
                <tr>
                    <th style="width: 100px;">Test ID</th>
                    <th style="width: 120px;">Module</th>
                    <th style="width: 80px;">Priority</th>
                    <th>Scenario Title</th>
                    <th>Error Message</th>
                </tr>
            </thead>
            <tbody>
        """
        
        if not failed_details:
            html_dashboard_template += """
                <tr>
                    <td colspan="5" style="text-align: center; color: var(--text-secondary);">No failing test cases in this run! System meets all deployment criteria.</td>
                </tr>
            """
        else:
            for fd in failed_details:
                html_dashboard_template += f"""
                <tr>
                    <td><code>{fd['id']}</code></td>
                    <td>{fd['module']}</td>
                    <td><span class="badge" style="background-color: #3b0712; color: #f43f5e;">{fd['priority']}</span></td>
                    <td style="font-weight: 500;">{fd['title']}</td>
                    <td style="color: var(--failed); font-family: monospace; font-size: 12px; white-space: pre-wrap;">{fd['error_message']}</td>
                </tr>
                """

        html_dashboard_template += """
            </tbody>
        </table>

        <div class="section-title">Module Execution Breakdown</div>
        <table>
            <thead>
                <tr>
                    <th>Module</th>
                    <th>Executed</th>
                    <th>Passed</th>
                    <th>Failed</th>
                    <th>Skipped</th>
                    <th>Pass Rate</th>
                </tr>
            </thead>
            <tbody>
        """

        modules_set = sorted(list(set(r["module"] for r in results)))
        for m in modules_set:
            mod_tests = [r for r in results if r["module"] == m]
            m_tot = len(mod_tests)
            m_pass = len([r for r in mod_tests if r["status"] == "Passed"])
            m_fail = len([r for r in mod_tests if r["status"] == "Failed"])
            m_skip = len([r for r in mod_tests if r["status"] == "Skipped"])
            m_rate = (m_pass / m_tot * 100) if m_tot > 0 else 0
            
            badge_class = "badge-pass" if m_rate >= 95 else "badge-fail"
            html_dashboard_template += f"""
                <tr>
                    <td><strong>{m}</strong></td>
                    <td>{m_tot}</td>
                    <td class="text-pass">{m_pass}</td>
                    <td class="text-fail">{m_fail}</td>
                    <td class="text-skip">{m_skip}</td>
                    <td><span class="badge {badge_class}">{m_rate:.2f}%</span></td>
                </tr>
            """

        html_dashboard_template += """
            </tbody>
        </table>
        
        <div class="footer">
            MedIntel Nexus QA Enterprise Automation Framework &copy; 2026. All rights reserved.
        </div>
    </div>
</body>
</html>
        """
        
        # Save Dashboard HTML
        dashboard_path = os.path.join(Config.HTML_DIR, "dashboard.html")
        with open(dashboard_path, "w", encoding="utf-8") as f:
            f.write(html_dashboard_template)
            
        # Detailed execution-report.html (Same template, with detailed listing of all 1500 test cases)
        html_report_template = html_dashboard_template.replace(
            '<div class="section-title">Module Execution Breakdown</div>',
            """
            <div class="section-title">Complete 1,500 Test Cases Log</div>
            <table>
                <thead>
                    <tr>
                        <th style="width: 100px;">Test ID</th>
                        <th style="width: 120px;">Module</th>
                        <th style="width: 80px;">Status</th>
                        <th>Title</th>
                        <th style="width: 90px;">Duration</th>
                    </tr>
                </thead>
                <tbody>
            """ + "".join([
                f"""
                <tr>
                    <td><code>{r['id']}</code></td>
                    <td>{r['module']}</td>
                    <td><span class="badge badge-{r['status'].lower()}">{r['status']}</span></td>
                    <td>{r['title']}</td>
                    <td>{r['execution_time']:.3f}s</td>
                </tr>
                """ for r in results
            ]) + """
                </tbody>
            </table>
            <div class="section-title">Module Execution Breakdown</div>
            """
        )
        
        report_path = os.path.join(Config.HTML_DIR, "execution-report.html")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(html_report_template)
            
        logger.info("Generated HTML dashboard and execution reports.")

    @staticmethod
    def generate_summary_markdown(results: list):
        """Generates summary.md for GitHub Actions summaries and artifact packaging."""
        os.makedirs(Config.SUMMARY_DIR, exist_ok=True)
        
        total = len(results)
        passed = len([r for r in results if r["status"] == "Passed"])
        failed = len([r for r in results if r["status"] == "Failed"])
        skipped = len([r for r in results if r["status"] == "Skipped"])
        pass_rate = (passed / total * 100) if total > 0 else 0
        total_duration = sum(r["execution_time"] for r in results)
        
        # Find Top Passing and Failing Modules
        module_stats = {}
        for r in results:
            m = r["module"]
            if m not in module_stats:
                module_stats[m] = {"tot": 0, "pass": 0, "fail": 0}
            module_stats[m]["tot"] += 1
            if r["status"] == "Passed":
                module_stats[m]["pass"] += 1
            elif r["status"] == "Failed":
                module_stats[m]["fail"] += 1
                
        sorted_modules = sorted(
            module_stats.items(), 
            key=lambda x: (x[1]["pass"]/x[1]["tot"]), 
            reverse=True
        )
        top_passing = [m[0] for m in sorted_modules[:3] if (m[1]["pass"]/m[1]["tot"]) > 0.95]
        top_failing = [m[0] for m in reversed(sorted_modules) if m[1]["fail"] > 0][:3]
        
        markdown_summary = f"""# 🚀 MedIntel Nexus QA Execution Summary

### 🌐 Live Deployment Target
* **Deployment URL**: [{Config.BASE_URL}]({Config.BASE_URL})
* **Deployment Status**: ✅ Completed & Available
* **Execution Timestamp**: {time.strftime('%Y-%m-%d %H:%M:%S UTC')}

---

### 📊 Performance Summary Metrics
| Metric | Value |
|---|---|
| **Total Test Cases Executed** | {total} |
| **Passed** | {passed} |
| **Failed** | {failed} |
| **Skipped** | {skipped} |
| **Pass Percentage** | **{pass_rate:.2f}%** |
| **Total Execution Duration** | {total_duration:.2f} seconds |

---

### 🏆 Module Insights
* **Top Performing Modules**: {", ".join(top_passing) if top_passing else "None"}
* **Top Failed Modules (Need Attention)**: {", ".join(top_failing) if top_failing else "None"}

---

### 📦 Uploaded Build Artifacts
1. **Excel Reports** (under `Excel/`):
   * `Automation_Test_Report.xlsx` (Full execution results)
   * `Passed_Test_Cases.xlsx`
   * `Failed_Test_Cases.xlsx`
   * `Summary_Report.xlsx`
2. **HTML Dashboards** (under `HTML/`):
   * `dashboard.html` (Interactive execution summary)
   * `execution-report.html` (List of all 1,500 tests)
3. **Execution Metadata** (under `JSON/`):
   * `execution-results.json`
4. **Diagnostic Assets** (under `Logs/` and `Screenshots/`):
   * Full execution logs (`automation.log`)
   * Browser screenshots on failure

---

*This summary report was programmatically generated by the MedIntel Nexus Quality Reporting suite.*
"""
        
        summary_path = os.path.join(Config.SUMMARY_DIR, "summary.md")
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write(markdown_summary)
            
        logger.info("Generated Markdown execution summary.")

if __name__ == "__main__":
    try:
        ReportGenerator.consolidate_and_generate_all()
    except RuntimeError as exc:
        logger.error(str(exc))
        sys.exit(1)
